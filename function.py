from numpy import empty
from openpyxl import load_workbook

def empty_row(workbook_path):
    book = load_workbook(workbook_path)
    ws = book.worksheets[0]
    return (ws.max_row + 1)